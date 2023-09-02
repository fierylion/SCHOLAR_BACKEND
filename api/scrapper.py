import queue
import threading
from typing import Any
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from pathlib import Path
import time
from selenium.webdriver.chrome.service import Service
from concurrent.futures import ThreadPoolExecutor
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
class Scrape():
    def __init__(self) -> None:
        print('Initializing scrapping!!!!!')
        self.options = Options()
        self.service = Service(ChromeDriverManager().install())
   
        self.options.add_argument("--incognito")

        self.options.add_argument("--start-maximized")

        self.options.add_argument('--disable-gpu')

        self.options.add_argument('--disable-dev-shm-usage')

        self.options.add_argument("--window-size=1920,1080")

        self.options.add_argument("--headless")

        self.options.add_argument('--disable-extensions')

        self.options.add_argument('--no-sandbox')

        self.options.add_argument('--remote-debugging-port=9222')
    
        path = Path('/usr/local/bin')
        self.path= path
        self.wb = openpyxl.Workbook()
        self.driver = webdriver.Chrome(options=self.options,service=self.service )
        # publications
        self.publication_results_queue = queue.Queue()

    def __call__(self, test_users, path) -> Any:
        for user in test_users:
            print(f'Going to next user: ${user}')
            self.driver.get(user)
            self.driver.implicitly_wait(2)
            user_details = {}
            try:
                user_details = self.fetch_user_details(self.driver)
                (ws, current_row) = self.create_excel(self.wb, user_details)
                publications = self.driver.find_elements(By.CSS_SELECTOR, '#gsc_a_b>tr')
                threads = []
                for (ind, pb) in enumerate(publications):
                    page_link = pb.find_element(By.TAG_NAME, 'a')
                    link = page_link.get_attribute('href')
                    thread = ThreadPoolExecutor().submit(self.fetch_publications, link)
                    threads.append(thread)
                    if (ind == len(publications) - 1):
                        try:
                            show_more = self.driver.find_element(By.ID, 'gsc_bpf_more')
                            show_more.click()
                            time.sleep(2)
                            new_publications = self.driver.find_elements(By.CSS_SELECTOR, '#gsc_a_b>tr')
                            for pb in new_publications:
                                if pb in publications:
                                    continue
                                else:
                                    publications.append(pb)
                        except Exception as e:
                            print(e)
                            break
                with ThreadPoolExecutor(max_workers=4) as executor:
                    for thread in threads:
                        executor.submit(thread.result)
                        
                self.fill_multiple_publications(ws, current_row)
            except Exception as e:
                print(e)
        self.wb.save(path)

    def fetch_publications(self, link):
    
        temp_driver = webdriver.Chrome(options=self.options,service=self.service)
        temp_driver.get(link)
        temp_driver.implicitly_wait(10)
        publications = {}

        def check_presence_of_from():
            try:
                from_ = temp_driver.find_element(By.ID, 'gsc_oci_title_gg').text
                return from_
            except Exception as e:
                return False

        try:
            publications['url'] = link
            pb = temp_driver.find_element(By.ID, "gsc_oci_title_wrapper")
            from_ = check_presence_of_from()
            publications['from_'] = from_ if from_ else ''
            publications['title'] = pb.find_element(By.ID, 'gsc_oci_title').text
            main_details = temp_driver.find_elements(By.CSS_SELECTOR, '#gsc_oci_table>div')
            main_details.pop()
            for detail in main_details:
                key = detail.find_element(By.CSS_SELECTOR, '.gsc_oci_field').text.lower()
                value = detail.find_element(By.CSS_SELECTOR, '.gsc_oci_value').text if key != 'total citations' else \
                detail.find_element(By.CSS_SELECTOR, '.gsc_oci_value a').text.split(' ')[2]
                publications[key] = value
            self.publication_results_queue.put(publications)
            temp_driver.close()
        except Exception as e:
            print(e)
            temp_driver.close()

    @staticmethod
    def fetch_user_details(dr):
        user_details = {}

        def details_table():
            table = dr.find_element(By.ID, 'gsc_rsb_st')
            rows = table.find_elements(By.TAG_NAME, 'tr')
            details = []
            for (ind, row) in enumerate(rows):
                th_or_tr = 'th' if ind == 0 else 'td'
                details.append([cell.text for cell in row.find_elements(By.TAG_NAME, th_or_tr)])
            return details

        try:
            time.sleep(2)
            basic_info = dr.find_element(By.ID, 'gsc_prf_i')
            user_details['name'] = basic_info.find_element(By.ID, 'gsc_prf_inw').text
            user_details['position'] = basic_info.find_element(By.CSS_SELECTOR, '.gsc_prf_il').text.split(',')[0]
            user_details['verification'] = basic_info.find_element(By.ID, 'gsc_prf_ivh').text
            user_details['fields'] = [fd.text for fd in basic_info.find_elements(By.CSS_SELECTOR, '#gsc_prf_int>a')]
            user_details['details'] = details_table()
            return user_details
        except Exception as e:
            print(e)

    def fill_multiple_publications(self, ws, current_row):
        ind = 1
        while not self.publication_results_queue.empty():
            pb = self.publication_results_queue.get()
            pb['no'] = ind
            current_row = self.publication_details_excel(ws, current_row, pb)
            ind += 1

    # write single publication to excel
    @staticmethod
    def publication_details_excel(ws, current_row, pb):
        # Expected document structure: {title: title, authors: authors, from_: from,  date: date, source: source, volume: volume, pages: pages, publisher: publisher, citations: citations, url: url}
        ws.cell(row=current_row, column=1).value = pb.get('no')
        ws.cell(row=current_row, column=2).value = pb.get('title', '')
        ws.cell(row=current_row, column=3).value = pb.get('authors', '')
        ws.cell(row=current_row, column=4).value = pb.get('from_', '')
        ws.cell(row=current_row, column=5).value = pb.get('publication date', '')
        ws.cell(row=current_row, column=6).value = pb.get('source', '')
        ws.cell(row=current_row, column=7).value = pb.get('volume', '')
        ws.cell(row=current_row, column=8).value = pb.get('pages', '')
        ws.cell(row=current_row, column=9).value = pb.get('publisher', '')
        ws.cell(row=current_row, column=10).value = pb.get('total citations', '')
        ws.cell(row=current_row, column=11).value = pb.get('conference', '')
        ws.cell(row=current_row, column=12).value = pb.get('url', '')
        return current_row + 1

    # create excel sheet and fill in user details
    @staticmethod
    def create_excel(wb, person_details):
        # document of structure {position: lecturer, name:baraka, verification:baraka, fields:[doctor, engineer, cousing], details:[[baraka, baraka, baraka, dafda], [baraka, baraka, baraka], [baraka, baraka, baraka]]}
        details = person_details.get('details')  # list of lists
        publication_fields = ['No', 'Title', 'From', 'Authors', 'Date', 'Source', 'Volume', 'Pages', 'Publisher',
                              'Citations', 'Conference', 'URL']
        ws = wb.create_sheet(f'Sheet-{person_details.get("name")}')
        ws['A1'] = 'Position'
        ws['B1'] = person_details.get('position')
        ws['A2'] = 'Name'
        ws['B2'] = person_details.get('name')

        ws['A3'] = 'Verification Details'
        ws['B3'] = person_details.get('verification')
        ws['A4'] = 'Fields'
        for (ind, fd) in enumerate(person_details.get('fields')):
            ws.cell(row=4, column=(ind + 2)).value = fd

        # fill in details table
        def fill_details_table():
            no_rows = len(details)
            for i in range(no_rows):
                for j in range(len(details[i])):
                    ws.cell(row=(i + 6), column=(j + 1)).value = details[i][j]

        fill_details_table()
        current_row = len(details) + 6
        ws.cell(row=current_row + 1, column=1).value = 'Publications'
        for (ind, fd) in enumerate(publication_fields):
            ws.cell(row=(current_row + 2), column=(ind + 1)).value = fd
        return (ws, current_row + 3)


